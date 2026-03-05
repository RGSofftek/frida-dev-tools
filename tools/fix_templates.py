import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency: openpyxl")
    print("Install it with: pip install openpyxl")
    sys.exit(1)


# Mandatory fields (as used in Actions.txt) mapped to
# (header text in Template row 1, default value to inject when empty)
MANDATORY_DEFAULTS: dict[str, tuple[str, str]] = {
    "ContractType": (
        "ContractType",
        "YP01",
    ),
    "OrgData": (
        "Org. Data",
        "CO03",
    ),
    "PurchasingGroup": (
        "Purchasing Group",
        "010",
    ),
    "PriceType": (
        "Price Type",
        "FLAT",
    ),
    "Quantity": (
        "Quantity",
        "420",
    ),
    "DeliveryPeriodFrom": (
        "Delivery Period From",
        "23.01.2026",
    ),
    "DeliveryPeriodTo": (
        "Delivery Period To",
        "27.01.2026",
    ),
    "BidPrice": (
        "BID Price",
        "1000",
    ),
    "ContractDate": (
        "Contract Date",
        "23.01.2026",
    ),
    "Plant": (
        "Plant",
        "CO9H",
    ),
    "Material": (
        "Material",
        "100126",
    ),
    "BP": (
        "BP",
        "7000003018",
    ),
}

RED_FILL = PatternFill(fill_type="solid", fgColor="FF0000")


def fix_header(excel_path: str, sheet_name: str = "Template") -> None:
    """
    - Ensure the header in row 1, column 10 is exactly 'Delivery Period To'.
    - Ensure all mandatory headers exist (adding columns if needed).
    - For each mandatory field, if a row value is empty, inject a default value and color the cell red.
    """
    excel_file = Path(excel_path)

    if not excel_file.exists():
        print(f"File not found: {excel_file}")
        return

    wb = openpyxl.load_workbook(excel_file)

    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found in {excel_file}")
        return

    ws = wb[sheet_name]

    # Ensure header in column 10 (J1) is 'Delivery Period To'
    ws.cell(row=1, column=10).value = "Delivery Period To"

    # Build mapping from header text to column index based on row 1
    header_to_col: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value:
            header_text = str(cell.value).strip()
            header_to_col[header_text] = cell.column

    # Ensure all mandatory headers exist; if missing, add them as new columns at the end
    max_col = ws.max_column
    for _, (header_text, _default_value) in MANDATORY_DEFAULTS.items():
        if header_text not in header_to_col:
            max_col += 1
            ws.cell(row=1, column=max_col).value = header_text
            header_to_col[header_text] = max_col
            print(f"Added missing header '{header_text}' as column {max_col}")

    # Determine last data row based on first empty in column A (ContractType)
    last_data_row = 1
    row_idx = 2
    while True:
        key_cell = ws.cell(row=row_idx, column=1)
        if key_cell.value in (None, ""):
            break
        last_data_row = row_idx
        row_idx += 1

    # For each data row up to last_data_row, inject defaults and color in red when a mandatory cell is empty.
    for _, (header_text, default_value) in MANDATORY_DEFAULTS.items():
        col_idx = header_to_col[header_text]
        for row_idx in range(2, last_data_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value in (None, ""):
                cell.value = default_value
                cell.fill = RED_FILL

    # Incoterms fix: any "DAP" should become "chaparral" and be colored red.
    incoterms_col = header_to_col.get("Incoterms")
    if incoterms_col is not None:
        for row_idx in range(2, last_data_row + 1):
            cell = ws.cell(row=row_idx, column=incoterms_col)
            if cell.value is None:
                continue
            text = str(cell.value).strip()
            if text.upper() == "DAP":
                cell.value = "chaparral"
                cell.fill = RED_FILL

    # Date fix: Delivery Period From/To must be future; if not, push them at least one month forward and color red.
    def _adjust_future_date(cell) -> None:
        raw = cell.value
        if raw is None or raw == "":
            return

        # Try to interpret as datetime first, then as string dd.MM.yyyy
        if isinstance(raw, datetime):
            dt = raw
        else:
            text = str(raw).strip()
            try:
                dt = datetime.strptime(text, "%d.%m.%Y")
            except ValueError:
                return

        today = datetime.today().date()
        d = dt.date()
        # If already in the future, leave as-is
        if d > today:
            return

        # Push at least one month forward; ensure strictly future
        new_date = d + timedelta(days=30)
        while new_date <= today:
            new_date = new_date + timedelta(days=30)

        cell.value = new_date.strftime("%d.%m.%Y")
        cell.fill = RED_FILL

    from_col = header_to_col.get("Delivery Period From")
    to_col = header_to_col.get("Delivery Period To")
    for row_idx in range(2, last_data_row + 1):
        if from_col is not None:
            _adjust_future_date(ws.cell(row=row_idx, column=from_col))
        if to_col is not None:
            _adjust_future_date(ws.cell(row=row_idx, column=to_col))

    wb.save(excel_file)
    print(f"Fixed headers and filled defaults in '{sheet_name}' of '{excel_file}'")


def main(argv: list[str]) -> int:
    """
    If called with:
      - no args: process all .xlsx files under 'Inputs' directory.
      - one arg: treat it as a directory and process all .xlsx files inside.
      - two args: <directory> <sheet_name>.
    """
    base_dir: Path
    sheet_name = "Template"

    if len(argv) == 1:
        base_dir = Path("Inputs")
    elif len(argv) == 2:
        base_dir = Path(argv[1])
    else:
        base_dir = Path(argv[1])
        sheet_name = argv[2]

    if not base_dir.exists() or not base_dir.is_dir():
        print(f"Directory not found or not a directory: {base_dir}")
        return 1

    excel_files = sorted(p for p in base_dir.glob("*.xlsx") if p.is_file())

    if not excel_files:
        print(f"No .xlsx files found in {base_dir}")
        return 0

    print(f"Processing {len(excel_files)} Excel file(s) in {base_dir} (sheet='{sheet_name}')")
    for excel_file in excel_files:
        fix_header(str(excel_file), sheet_name)

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))

